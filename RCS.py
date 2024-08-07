import openpyxl
import pymysql
from tkinter import *
from tkinter.ttk import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl.styles import PatternFill
import pyperclip
import os
from tkinter.filedialog import askopenfilename
import customtkinter as ctk
from PIL import Image, ImageTk
import cryptography
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")
host = '192.168.1.3'  # Replace with the IP address or hostname of the MySQL server
# user = 'MicrosoftAcount/administrator'
# password = 'Aa123456'
database = 'request_control'

conn = pymysql.connect(
    host= host ,user='root', password='1122@Nilab8711', database='request_control'
)
cursor = conn.cursor()


def upload():
    def errorWindow(error_massage):
        newWindow = ctk.CTk()
        newWindow.title("Error")
        newWindow.geometry("400x100")
        title_label = ctk.CTkLabel(newWindow, text=error_massage)
        title_label.place(x=80, y=10)
        close_button = ctk.CTkButton(newWindow, text="ok", command=newWindow.destroy)
        close_button.place(x=110, y=60)
        newWindow.mainloop()

    def run_program():
        window.destroy()
        View()

    def chooseTable(name):
        return {
            "فاضلاب قم 5 ساله": "quem",
            "فاضلاب خین عرب": "khin",
            "فاضلاب التیمور": "alteymour",
            "آبرسانی جاسک": "jask",
            "رودان 2": "roudan",
            "رشت": "rasht",
            "مرکزی": "markazi",
            "زاهدان": "zahedan",
        }[name]

    def chooseCode(name):
        return {
            "زاهدان": 900,
            "فاضلاب قم 5 ساله": 777,
            "فاضلاب خین عرب": 770,
            "فاضلاب التیمور": 880,
            "آبرسانی جاسک": 667,
            "رودان 2": 666,
            "رشت": 210,
            "مرکزی": 110,
        }[name]

    def checkForFile(fileLocation):
        file_name = os.path.basename(fileLocation).replace(".xlsm", "")
        query = "SELECT * FROM main WHERE req_n = %s"
        value = (file_name,)
        cursor.execute(query, value)
        if cursor.fetchone():
            updateWindow(file_name, fileLocation)
        else:
            save_data_to_database(file_name, False, fileLocation)

    def updateWindow(file_name, file_path):
        newWindow = ctk.CTk()
        newWindow.title("Error")
        newWindow.geometry("400x100")
        title_label = ctk.CTkLabel(
            newWindow,
            text="کنید؟ آپدیت را آن میخواهید آیا. دارد وجود دیتابیس در فایل این",
        )
        title_label.place(x=60, y=10)

        def update():
            newWindow.destroy()
            save_data_to_database(file_name, True, file_path)

        update_button = ctk.CTkButton(newWindow, text="update", command=update)
        update_button.place(x=220, y=60)

        def close():
            newWindow.destroy()

        close_button = ctk.CTkButton(newWindow, text="cancle", command=close)
        close_button.place(x=50, y=60)
        newWindow.mainloop()

    def save_data_to_database(file_name, delete_needed, file_path):
        try:
            workbook = openpyxl.load_workbook(file_path, keep_vba=True, data_only=True)
            worksheet = workbook[file_name.replace(".xlsm", "")]
            req_n = file_name
            # request type
            if "REM" in req_n:
                req_t = "کالا"
            else:
                req_t = "کار"
            # order date
            o_date = worksheet["N6"].value
            # referral date
            if req_t == "کالا":
                ref_date = worksheet["N17"].value
            else:
                ref_date = worksheet["N16"].value
            # request status
            if worksheet["X8"].value == True:
                status = "توقف"
            elif (
                worksheet["X7"].value != True
                or worksheet["N17"].value == "XXXX/XX/XX"
                or worksheet["N16"].value == "XXXX/XX/XX"
            ):
                errorWindow(req_n + " فایل برای پروژه مدیر تایید عدم ")
                status = "invalid"
            else:
                if worksheet["X13"].value == TRUE:
                    if req_t == "کالا":
                        status = "تکمیل و ارسال به سایت"
                    else:
                        status = "تکمیل کار"
                elif worksheet["X12"].value == TRUE:
                    if req_t == "کالا":
                        status = "پروسه ی خرید"
                    else:
                        status = "پروسه ی ارجاع کار به پیمانکار"

                elif worksheet["X11"].value == TRUE:
                    status = "تأمین بودجه"
                elif worksheet["X10"].value == TRUE:
                    if req_t == "کالا":
                        status = "جستجوی کالا "
                    else:
                        status = "جستجوی پیمانکار"
                elif worksheet["X10"].value == TRUE:
                    status = "درخواست اطلاعات از سایت"
                else:
                    status = "ارجا به بازرگانی"
            # project name
            project_name = worksheet["G6"].value

            # get the project name and code
            table = chooseTable(project_name)
            code = chooseCode(project_name)

            # check if it needs updates
            if delete_needed:
                query = f"DELETE FROM {table} WHERE req_n=%s"
                value = (file_name,)
                query_main = f"DELETE FROM main WHERE req_n=%s"
                value_main = (file_name,)
                cursor.execute(query, value)
                cursor.execute(query_main, value_main)
            # applicant
            Applicant = worksheet["D6"].value
            # get the materials value
            for row in range(8, 14):
                if req_t == "کالا":
                    prod = worksheet.cell(row=row, column=3).value
                    qty = worksheet.cell(row=row, column=9).value
                    available = worksheet.cell(row=row, column=11).value
                    place_of_usage = worksheet.cell(row=row, column=14).value
                    unit = worksheet.cell(row=row, column=13).value
                else:
                    prod = worksheet.cell(row=row, column=3).value
                    place_of_usage = worksheet["K7"].value
                    qty = None
                    available = None
                    unit = None
                if prod != None and prod != "شرح خدمات درخواستی :":
                    sql = f"INSERT INTO {table} (product, req_n, Qty, o_date, available_in_stock, ref_date, place_of_usage, unit ,Applicant, st_request, req_t) VALUES (%s, %s, %s,%s,%s, %s, %s, %s, %s, %s, %s)"
                    val = (
                        prod,
                        req_n,
                        qty,
                        o_date,
                        available,
                        ref_date,
                        place_of_usage,
                        unit,
                        Applicant,
                        status,
                        req_t,
                    )
                    cursor.execute(sql, val)
            if status != "invalid":
                sql_main = f"INSERT INTO main (project, req_n, o_date, f_date, p_code, st_request, req_t, location) VALUES (%s ,%s, %s, %s, %s, %s, %s, %s)"
                val_main = (
                    project_name,
                    req_n,
                    o_date,
                    ref_date,
                    code,
                    status,
                    req_t,
                    file_path,
                )
                cursor.execute(sql_main, val_main)
                conn.commit()
                done_wondow()

        except KeyError as e:
            errorWindow(
                "است قبول قابل غیر فایل نام\n است RE#-#####-###-### قبول قابل فرمت"
            )

    def done_wondow():
        new_window = ctk.CTk()
        new_window.title("saved")
        new_window.geometry("300x100")
        done_label = ctk.CTkLabel(
            new_window,
            text="شد ثبت موفقیت با اطلاعات",
        )
        done_label.place(x=100, y=15)
        ok_button = ctk.CTkButton(new_window, text="ok", command=new_window.destroy)
        ok_button.place(x=80, y=60)
        new_window.mainloop()

    def codeTableDisplay():
        table = ttk.Treeview(window, columns=("1", "2"), show="headings", height=8)
        table.pack()
        table.column("1", anchor=CENTER, stretch=YES, width=100)
        table.heading("1", text="نام پروژه")
        table.column("2", anchor=CENTER, stretch=YES, width=50)
        table.heading("2", text="کد پروژه")
        table.insert("", "end", values=("فاضلاب قم 5 ساله", "777"))
        table.insert("", "end", values=("فاضلاب خین عرب", "770"))
        table.insert("", "end", values=("فاضلاب التیمور", "880"))
        table.insert("", "end", values=("آبرسانی جاسک", "667"))
        table.insert("", "end", values=("رودان 2", "666"))
        table.insert("", "end", values=("رشت", "210"))
        table.insert("", "end", values=("مرکزی", "110"))
        table.insert("", "end", values=("زاهدان", "900"))

        table.place(x=250, y=100)  # change


    # Create the main window
    window = ctk.CTk()
    window.geometry("500x500")
    window.title("سیستم کنترل درخواست کالا و کار")
    title_label = ctk.CTkLabel(window, text="البرز صنعت نیلآب شرکت")
    title_label.place(x=200, y=10)

    # Create the project name label and entry widgets
    request_number_label = ctk.CTkLabel(window, text="Select File : ")
    request_number_label.place(x=10, y=80)
    find_button = ctk.CTkButton(window, text="  bowers   ", command=lambda:checkForFile(askopenfilename().replace("//SERVER/Department/Archive","Z:")))
    find_button.place(x=90, y=78)

    back_button = ctk.CTkButton(
        window, text="   back   ", command=lambda: run_program()
    )
    back_button.place(x=10, y=125)
    # Display the company logo image
    image = Image.open("Z:/Archive/1. Projects/RCS-v01/root/logo5.png")  # change
    resized_image = image.resize((100, 80))
    photo = ImageTk.PhotoImage(resized_image)
    canvas = Canvas(
        window, bg="#282424", width=resized_image.width, height=resized_image.height
    )
    canvas.pack(side="top", anchor="ne")
    canvas.create_image(0, 0, anchor=NW, image=photo)

    codeTableDisplay()
    # Run the GUI
    window.mainloop()
    cursor.close()
    conn.close()


def View():
    def reverse_name(name):
        return {
            "ساله 5 قم فاضلاب": "فاضلاب قم 5 ساله",
            "عرب خین فاضلاب": "فاضلاب خین عرب",
            "التیمور فاضلاب": "فاضلاب التیمور",
            "جاسک آبرسانی": "آبرسانی جاسک",
            "2 رودان": "رودان 2",
            "رشت": "رشت",
            "مرکزی": "مرکزی",
            "زاهدان":"زاهدان",
        }[name]

    def errorwindow(string, n):
        new_window = ctk.CTk()
        new_window.title("Error")
        new_window.geometry("300x100")
        title_label = ctk.CTkLabel(
            new_window,
            text=string,
        )
        title_label.pack()

        if n == 1:
            title_label.place(x=100, y=10)
        elif n == 2:
            title_label.place(x=70, y=10)
        else:
            title_label.place(x=70, y=10)

        update_button = ctk.CTkButton(new_window, text="ok", command=new_window.destroy)
        update_button.place(x=80, y=60)
        new_window.mainloop()

    def checkForFile(search, table_name):
        if len(search) == 0:
            querynull = f"SELECT * FROM {table_name}"
            cursor.execute(querynull)
        else:
            query = f"SELECT * FROM {table_name} WHERE product LIKE '%{search}%' OR req_n LIKE '%{search}%' "
            cursor.execute(query)
        data = cursor.fetchall()
        if table_name == "main":
            mainTable(data)
            mainFiltering(data)
        else:
            showTable(data, table_name)

    def mainFiltering(data):
        button = ctk.CTkButton(window, text="Search in main", command=filtering)
        button.place(relx=0.51, rely=0.09)

    def printData(data, table_name):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        if table_name == "main":
            worksheet.cell(row=1, column=1).value = "نام پروژه"
            worksheet.cell(row=1, column=2).value = "کد پروژه"
            worksheet.cell(row=1, column=3).value = "شماره درخواست"
            worksheet.cell(row=1, column=4).value = "تاریخ درخواست"
            worksheet.cell(row=1, column=5).value = "تاریخ ارجا"
            worksheet.cell(row=1, column=6).value = "نوع درخواست"
            worksheet.cell(row=1, column=7).value = "وضعیت درخواست"
            worksheet.cell(row=1, column=8).value = "پرداخت اول"
            worksheet.cell(row=1, column=9).value = "پرداخت دوم"
            worksheet.cell(row=1, column=10).value = "توضیحات"
        else:
            worksheet.cell(row=1, column=1).value = "درخواست"
            worksheet.cell(row=1, column=2).value = "شماره درخواست"
            worksheet.cell(row=1, column=3).value = "تاریخ درخواست"
            worksheet.cell(row=1, column=4).value = "تاریخ ارجا"
            worksheet.cell(row=1, column=5).value = "تعداد"
            worksheet.cell(row=1, column=6).value = "موجود"
            worksheet.cell(row=1, column=7).value = "واحد"
            worksheet.cell(row=1, column=8).value = "محل مصرف"
            worksheet.cell(row=1, column=9).value = "درخواست کننده"
            worksheet.cell(row=1, column=10).value = "وضعیت درخواست"
            worksheet.cell(row=1, column=11).value = "نوع درخواست"
            worksheet.cell(row=1, column=12).value = "توضیحات"

        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx).value = cell_value
                if cell_value == "توقف":
                    # Apply red fill to the entire row
                    fill = PatternFill(
                        start_color="FF0000", end_color="FF0000", fill_type="solid"
                    )
                    for cell in worksheet[row_idx]:
                        cell.fill = fill
                elif cell_value == "تکمیل و ارسال به سایت":
                    # Apply red fill to the entire row
                    fill = PatternFill(
                        start_color="00FF50", end_color="00FF50", fill_type="solid"
                    )
                    for cell in worksheet[row_idx]:
                        cell.fill = fill
        for column in range(ord("A"), ord("M")):
            column_letter = chr(column)
            worksheet.column_dimensions[column_letter].width = 18

        workbook.save(file_path)
        os.startfile(file_path)

        newWindow = ctk.CTk()

        newWindow.geometry("300x100")
        title_label = ctk.CTkLabel(
            newWindow,
            text="شد ثبت موفقیت با اطلاعات",
        )
        title_label.place(x=100, y=10)
        close_button = ctk.CTkButton(newWindow, text="ok", command=newWindow.destroy)
        close_button.place(x=80, y=60)
        newWindow.mainloop()

    def copy_to_clipboard(event, table):
        selected_row = table.focus()
        data = table.item(selected_row)["values"]
        if event == "table":
            pyperclip.copy(data[2])
        else:
            pyperclip.copy(data[3])

    def open_file(table):
        selected_row = table.focus()
        data = table.item(selected_row)["values"]
        query = f"SELECT * FROM main WHERE req_n = %s"
        value = (data[3],)
        cursor.execute(query, value)
        row = cursor.fetchall()
        try:
            os.startfile(row[0][11])
        except (FileNotFoundError, TypeError):
            errorwindow("نشد یافت فایل \n کنید آپدیت را نظر مورد فایل لطفا", 2)

    def handle_double_click(table):
        selected_row = table.focus()
        data = table.item(selected_row)["values"]
        req_n = data[3]
        newWindow1 = ctk.CTk()
        newWindow1.title(req_n)
        newWindow1.geometry("400x100")
        title_label = ctk.CTkLabel(
            newWindow1,
            text=req_n,
        )
        title_label.place(x=150, y=10)

        def comment():
            newWindow1.destroy()
            newWindow = ctk.CTk()
            newWindow.geometry("400x130")
            title_label = ctk.CTkLabel(
                newWindow,
                text=" سفارش  توضیحات ",
            )
            title_label.place(x=165, y=5)
            comment_entry = ctk.CTkEntry(
                newWindow, placeholder_text="comment", width=300
            )
            comment_entry.place(x=50, y=40)

            def insertButton():
                sql = f"Update main set comment = %s where req_n = %s"
                val = (comment_entry.get(), req_n)
                cursor.execute(sql, val)
                conn.commit()
                newWindow.destroy()

            insert_button = ctk.CTkButton(newWindow, text="ثبت", command=insertButton)
            insert_button.place(x=140, y=85)
            newWindow.mainloop()

        comment_button = ctk.CTkButton(
            newWindow1, text="توضیحات کردن اضافه", command=comment
        )
        comment_button.place(x=40, y=40)

        def pay_window():
            def insert():
                try:
                    save = True
                    workbook = openpyxl.load_workbook(
                        "Z:/Archive/1. Projects/RCS-v01/root/payment order.xlsx"  # change
                    )
                    sheet = workbook["Sheet1"]
                    amount = amount_entry.get()
                    sheet["F6"] = req_n
                    sheet["E10"] = receiver_entry.get()
                    sheet["L3"] = number_entry.get()
                    sheet["K9"] = amount
                    if data[8] == "None":
                        sql_main = f"Update main set payment1 = %s where req_n = %s"
                        val_main = (amount, req_n)
                        cursor.execute(sql_main, val_main)
                        conn.commit()
                    elif data[9] == "None":
                        sql_main = f"Update main set payment2 = %s where req_n = %s"
                        val_main = (amount, req_n)
                        cursor.execute(sql_main, val_main)
                        conn.commit()
                    else:
                        errorwindow("است شده پر درخواست 2 تعداد", 3)
                        save = False
                    if save == True:
                        path = filedialog.asksaveasfilename(
                            defaultextension=".xlsx", initialfile=req_n + "-PO"
                        )
                        workbook.save(path)
                        os.startfile(path)
                except PermissionError as e:
                    errorwindow("نمیشود داده باز فایل برای دسترسی اجازه", 2)
                newWindow.destroy()

            newWindow1.destroy()
            newWindow = ctk.CTk()
            newWindow.geometry("400x200")
            title_label = ctk.CTkLabel(
                newWindow,
                text=req_n + " پرداخت دستور صدور ",
            )
            title_label.place(x=80, y=10)
            receiver_label = ctk.CTkLabel(newWindow, text=": کننده دریافت")
            receiver_label.place(x=290, y=40)
            receiver_entry = ctk.CTkEntry(newWindow, placeholder_text="اداری")
            receiver_entry.place(x=110, y=45)

            amount_label = ctk.CTkLabel(newWindow, text=": مبلغ")
            amount_label.place(x=320, y=70)
            amount_entry = ctk.CTkEntry(newWindow, placeholder_text="2000000")
            amount_entry.place(x=110, y=75)

            number_entry = ctk.CTkEntry(newWindow, placeholder_text="103")
            number_entry.place(x=110, y=105)
            number_label = ctk.CTkLabel(newWindow, text=": شماره")
            number_label.place(x=320, y=100)

            insert_button = ctk.CTkButton(newWindow, text="insert", command=insert)
            insert_button.place(x=210, y=140)

            close_button = ctk.CTkButton(
                newWindow, text="cancle", command=newWindow.destroy
            )
            close_button.place(x=40, y=140)
            newWindow.mainloop()

        payment_button = ctk.CTkButton(
            newWindow1, text="پرداخت دستور صدور", command=pay_window
        )
        payment_button.place(x=220, y=40)

        # def change_statuse():
        #     newWindow1.destroy()
        #     status_window = ctk.CTk()
        #     status_window.geometry("400x200")
        #     title_label = ctk.CTkLabel(
        #         status_window,
        #         text=": جدید وضعیت",
        #     )
        #     title_label.place(x=300, y=15)
        #     status = ctk.StringVar()
        #     if data[6] == "کالا":
        #         options = ctk.CTkOptionMenu(
        #             status_window,
        #             variable=status,
        #             values=(
        #                 "سایت از اطلاع درخواست",
        #                 "کالا جو و جست",
        #                 "بودجه تامین",
        #                 "خرید پورسه",
        #                 "سایت به ارسال و تکمیل",
        #             ),
        #         )
        #     else:
        #         options = ctk.CTkOptionMenu(
        #             status_window,
        #             variable=status,
        #             values=(
        #                 "سایت از اطلاع درخواست",
        #                 "پیمنکار جست و جو",
        #                 "بودجه تامین",
        #                 "پیمانکار به کار ارجاع پروسه",
        #                 "کار تکمیل",
        #             ),
        #         )
        #     options.place(x=30, y=10)
        #     def reverse(change):
        #         return {
        #             "سایت از اطلاع درخواست":"درخواست اطلاع از سایت",
        #             "کالا جو و جست":"جستجوی کالا",
        #             "بودجه تامین":"تأمین بودجه",
        #             "خرید پورسه":"پروسه ی خرید",
        #             "سایت به ارسال و تکمیل":"تکمیل و ارسال به سایت",
        #             "پیمنکار جست و جو":"جستجوی پیمانکار",
        #             "پیمانکار به کار ارجاع پروسه":"پروسه ی ارجاع کار به پیمانکار",
        #             "کار تکمیل":"تکمیل کار",
        #         } [change]
        #     def changes_to_excel(change):
        #         query = "SELECT * FROM main WHERE req_n = %s"
        #         value = (req_n,)
        #         cursor.execute(query, value)
        #         data = cursor.fetchone()
        #         try:
        #             workbook = openpyxl.load_workbook(data[11])
        #             # workbook.security.lockStructure = False
        #             sheet = workbook.active
        #             if change == "تکمیل و ارسال به سایت" or change == "تکمیل کار":
        #                 sheet["X13"] = True
        #             elif change == "پروسه ی ارجاع کار به پیمانکار" or change == "پروسه ی خرید":
        #                 sheet["X12"] = True
        #             elif change == "تأمین بودجه":
        #                 sheet["X11"] = True
        #             elif change == "جستجوی پیمانکار" or change == "جستجوی کالا ":
        #                 sheet["X10"] = True
        #             else:
        #                 sheet["X9"] = True
        #             print(data[11].rsplit('.', 1)[0] + '.xlsx')
        #             workbook.save(data[11].rsplit('.', 1)[0] + '.xls')
        #         except FileExistsError:
        #             errorwindow("است اشتباه فایل ادرس", 2)
        #         sql = f"Update main set st_request = %s where req_n = %s"
        #         val = (change, req_n)
        #         cursor.execute(sql, val)
        #         conn.commit()
        #         status_window.destroy()

        #     insert_button = ctk.CTkButton(
        #         status_window, text="ثبت", command=lambda:changes_to_excel(reverse(status.get()))
        #     )
        #     insert_button.place(x=225, y=80)
        #     status_window.mainloop()
                
        # status_button = ctk.CTkButton(
        #     newWindow1, text="وضعیت تغییر", command=change_statuse
        # )
        # status_button.place(x=210, y=40)

        newWindow1.mainloop()
        

    def mainTable(data):
        def print():
            printData(list(data), "main")
    
        print_button = ctk.CTkButton(window, text="   print   ", command=print)
        print_button.place(relx=0.87, rely=0.25)
        table = ttk.Treeview(
            window,
            columns=("1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11"),
            show="headings",
            height=20,
        )
        table.pack()
        table.column("1", anchor=CENTER, stretch=YES, width=40)
        table.heading("1", text="Id")
        table.column("2", anchor=CENTER, stretch=YES, width=125)
        table.heading("2", text="نام پروژه")
        table.column("3", anchor=CENTER, stretch=YES, width=50)
        table.heading("3", text="کد پروژه")
        table.column("4", anchor=CENTER, stretch=YES, width=145)
        table.heading("4", text="شماره درخواست")
        table.column("5", anchor=CENTER, stretch=YES, width=100)
        table.heading("5", text="تاریخ درخواست")
        table.column("6", anchor=CENTER, stretch=YES, width=85)
        table.heading("6", text="تاریخ ارجا")
        table.column("7", anchor=CENTER, stretch=YES, width=75)
        table.heading("7", text="نوع درخواست")
        table.column("8", anchor=CENTER, stretch=YES, width=160)
        table.heading("8", text="وضعیت درخواست")
        table.column("9", anchor=CENTER, stretch=YES, width=100)
        table.heading("9", text="پرداخت 1")
        table.column("10", anchor=CENTER, stretch=YES, width=100)
        table.heading("10", text="پرداخت 2")
        table.column("11", anchor=CENTER, stretch=YES, width=300)
        table.heading("11", text="توضیحات")
        table.place(relx=0.03, rely=0.3)  # change
        i = 0
        for row in data:
            table.insert(
                "",
                "end",
                text=i,
                values=(
                    i + 1,
                    data[i][0],
                    data[i][1],
                    data[i][2],
                    data[i][3],
                    data[i][4],
                    data[i][5],
                    data[i][6],
                    data[i][7],
                    data[i][8],
                    data[i][9],
                ),
            )
            i += 1
        table.bind("<Double-1>", lambda event: handle_double_click(table))
        table.bind("<c>", lambda event: copy_to_clipboard("main", table))
        table.bind("<Return>", lambda event: open_file(table))

    def showTable(data, table_name):
        def print():
            printData(list(data), table_name)

        print_button = ctk.CTkButton(window, text=" print ", command=print)
        print_button.place(relx=0.87, rely=0.25)
        table = ttk.Treeview(
            window,
            columns=(
                "1",
                "2",
                "3",
                "4",
                "5",
                "6",
                "7",
                "8",
                "9",
                "10",
                "11",
                "12",
                "13",
            ),
            show="headings",
            height=20,
        )
        table.pack()
        table.column("1", anchor=CENTER, stretch=YES, width=10)
        table.heading("1", text="Id")
        table.column("2", anchor=CENTER, stretch=YES, width=250)
        table.heading("2", text="درخواست")
        table.column("3", anchor=CENTER, stretch=YES, width=115)
        table.heading("3", text="شماره درخواست")
        table.column("4", anchor=CENTER, stretch=YES, width=85)
        table.heading("4", text="تاریخ درخواست")
        table.column("5", anchor=CENTER, stretch=YES, width=70)
        table.heading("5", text="تاریخ ارجا")
        table.column("10", anchor=CENTER, stretch=YES, width=100)
        table.heading("10", text="درخواست کننده")
        table.column("6", anchor=CENTER, stretch=YES, width=35)
        table.heading("6", text="تعداد")
        table.column("7", anchor=CENTER, stretch=YES, width=40)
        table.heading("7", text="موجود")
        table.column("9", anchor=CENTER, stretch=YES, width=120)
        table.heading("9", text="محل مصرف")
        table.column("8", anchor=CENTER, stretch=YES, width=50)
        table.heading("8", text="واحد")
        table.column("11", anchor=CENTER, stretch=YES, width=130)
        table.heading("11", text="وضعیت درخواست")
        table.column("12", anchor=CENTER, stretch=YES, width=75)
        table.heading("12", text="نوع درخواست")
        table.column("13", anchor=CENTER, stretch=YES, width=200)
        table.heading("13", text="توضیحات")
        table.place(relx=0.03, rely=0.3)
        i = 0
        for row in data:
            table.insert(
                "",
                "end",
                text=i,
                values=(
                    i + 1,
                    data[i][0],
                    data[i][1],
                    data[i][2],
                    data[i][3],
                    data[i][4],
                    data[i][5],
                    data[i][6],
                    data[i][7],
                    data[i][8],
                    data[i][9],
                    data[i][10],
                    data[i][11],
                ),
                tag=("odd"),
            )
            i += 1
        table.bind("<c>", lambda event: copy_to_clipboard("table", table))

        def comment():
            selected_row = table.focus()
            row = table.item(selected_row)["values"]
            product = row[1]
            newWindow = ctk.CTk()
            newWindow.geometry("400x130")
            title_label = ctk.CTkLabel(
                newWindow,
                text=" سفارش  توضیحات ",
            )
            title_label.place(x=165, y=5)
            comment_entry = ctk.CTkEntry(
                newWindow, placeholder_text="comment", width=300
            )
            comment_entry.place(x=50, y=40)

            def insertButton():
                sql = f"Update {table_name} set comment = %s where product = %s"
                val = (comment_entry.get(), product)
                cursor.execute(sql, val)
                conn.commit()
                newWindow.destroy()

            insert_button = ctk.CTkButton(newWindow, text="ثبت", command=insertButton)
            insert_button.place(x=140, y=85)
            newWindow.mainloop()

        table.bind("<Double-1>", lambda event: comment())

    def chooseTable(name):
        return {
            "ساله 5 قم فاضلاب": "quem",
            "عرب خین فاضلاب": "khin",
            "التیمور فاضلاب": "alteymour",
            "جاسک آبرسانی": "jask",
            "2 رودان": "roudan",
            "رشت": "rasht",
            "مرکزی": "markazi",
            "main": "main",
            "زاهدان":"zahedan",
        }[name]

    def discriptions():
        table = ttk.Treeview(window, columns=("1", "2"), show="headings", height=8)
        table.pack()
        table.column("1", anchor=CENTER, stretch=YES, width=100)
        table.heading("1", text="نام پروژه")
        table.column("2", anchor=CENTER, stretch=YES, width=50)
        table.heading("2", text="کد پروژه")
        table.insert("", "end", values=("فاضلاب قم 5 ساله", "777"))
        table.insert("", "end", values=("فاضلاب خین عرب", "770"))
        table.insert("", "end", values=("فاضلاب التیمور", "880"))
        table.insert("", "end", values=("آبرسانی جاسک", "667"))
        table.insert("", "end", values=("رودان 2", "666"))
        table.insert("", "end", values=("رشت", "210"))
        table.insert("", "end", values=("مرکزی", "110"))
        table.insert("", "end", values=("زاهدان", "900"))

        table.place(relx=0.68, rely=0.02)  # change

    # Create the main window
    window = ctk.CTk()
    window.title("سیستم کنترل درخواست کالا و کار")
    # window.geometry("1200x500")
    window.after(0, lambda: window.wm_state('zoomed'))
    image = Image.open("Z:/Archive/1. Projects/RCS-v01/root/logo5.png")  # change
    resized_image = image.resize((140, 120))
    photo = ImageTk.PhotoImage(resized_image)
    canvas = Canvas(
        window, bg="#282424", width=resized_image.width, height=resized_image.height
    )
    canvas.pack(side="top", anchor="ne")
    canvas.create_image(0, 0, anchor=NW, image=photo)

    #####################################################
    request_number_label = ctk.CTkLabel(window, text="request num or product:")
    request_number_label.place(relx=0.035, rely=0.03)
    # Create the request number text box
    request_number_entry = ctk.CTkEntry(window)
    request_number_entry.place(relx=0.15, rely=0.03)
    ######################################################
    name_label = ctk.CTkLabel(window, text="البرز صنعت نیلآب شرکت")
    name_label.place(relx=0.4, rely=0.01)
    #####################################################
    # Create a label for the project name choice box
    project_name_label = ctk.CTkLabel(window, text="Project Name:")
    project_name_label.place(relx=0.035, rely=0.09)

    # Create the project name choice box
    project_name = ctk.StringVar()
    options = ctk.CTkOptionMenu(
        window,
        variable=project_name,
        values=(
            "مرکزی",
            "رشت",
            "2 رودان",
            "جاسک آبرسانی",
            "التیمور فاضلاب",
            "عرب خین فاضلاب",
            "ساله 5 قم فاضلاب",
            "زاهدان",
        ),
    )
    options.place(relx=0.39, rely=0.09)

    project_name_var = ctk.StringVar()
    options = ctk.CTkOptionMenu(
        window,
        variable=project_name_var,
        values=(
            "مرکزی",
            "زاهدان",
            "رشت",
            "2 رودان",
            "جاسک آبرسانی",
            "التیمور فاضلاب",
            "عرب خین فاضلاب",
            "ساله 5 قم فاضلاب",
            "main",
        ),
    )
    project_name_var.set("main")
    options.place(relx=0.12, rely=0.09)

    def search_project():
        project_name.set("")
        selected_value = project_name_var.get()
        checkForFile(request_number_entry.get(), chooseTable(selected_value))

    project_name_button = ctk.CTkButton(window, text="Search", command=search_project)
    project_name_button.place(relx=0.24, rely=0.09)

    ###################################################

    def filtering():
        table_name = reverse_name(project_name.get())
        query = f"SELECT * FROM main WHERE project = %s"
        value = (table_name,)
        cursor.execute(query, value)
        data = cursor.fetchall()
        project_name_var.set("main")
        mainTable(data)

    button = ctk.CTkButton(window, text="Search in main", command=filtering)
    button.place(relx=0.51, rely=0.09)

    def uploadB():
        window.destroy()
        upload()

    upload_button = ctk.CTkButton(window, text="Upload New File", command=uploadB)
    upload_button.place(relx=0.035, rely=0.16)

    discriptions()
    # Run the GUI
    window.mainloop()


####### main
View()
