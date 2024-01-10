import sys
import os
import openpyxl
import pymysql
import openpyxl
from tkinter import *
from tkinter.ttk import *
import customtkinter as ctk


class project_manager(Exception):
    pass

conn = pymysql.connect(
    host= '192.168.1.3' ,user='root', password='1122@Nilab8711', database='request_control'
)
cursor = conn.cursor()


def errorWindow(error_massage):
    newWindow = ctk.CTk()
    newWindow.title("Error")
    newWindow.geometry("400x100")
    title_label = ctk.CTkLabel(newWindow, text=error_massage)
    title_label.place(x=80, y=10)
    close_button = ctk.CTkButton(newWindow, text="ok", command=newWindow.destroy)
    close_button.place(x=110, y=60)
    newWindow.mainloop()

def chooseTable(name):
        return {
            "فاضلاب قم 5 ساله": "quem",
            "فاضلاب خین عرب": "khin",
            "فاضلاب التیمور": "alteymour",
            "آبرسانی جاسک": "jask",
            "رودان 2": "roudan",
            "رشت": "rasht",
            "مرکزی": "markazi",
            "زاهدان":"zahedan",
        }[name]
def save_data_to_database(delete_needed, file_path):
    try:
        req_n = os.path.basename(file_path).replace(".xlsm", "")
        workbook = openpyxl.load_workbook(file_path, keep_vba=True, data_only=True)
        worksheet = workbook[req_n.replace(".xlsm", "")]
        
        # request type
        if "REM" in req_n:
            req_t = "کالا"
        else:
            req_t = "کار"
        # order date
        o_date = worksheet["N6"].value
        # referral date
        if req_t == "کالا":
            ref_date = worksheet["N17"].value
        else:
            ref_date = worksheet["N16"].value
        # request status
        if worksheet["X8"].value == True:
            status = "توقف"
        elif (
            worksheet["X7"].value == FALSE
            or worksheet["N17"].value == "XXXX/XX/XX"
            or worksheet["N16"].value == "XXXX/XX/XX"
        ):
            status = "invalid"
            raise project_manager(" فایل برای پروژه مدیر تایید عدم ")
            

        if worksheet["X13"].value == TRUE:
            if req_t == "کالا":
                status = "تکمیل و ارسال به سایت"
            else:
                status = "تکمیل کار"
        elif worksheet["X12"].value == TRUE:
            if req_t == "کالا":
                status = "پروسه ی خرید"
            else:
                status = "پروسه ی ارجاع کار به پیمانکار"

        elif worksheet["X11"].value == TRUE:
            status = "تأمین بودجه"
        elif worksheet["X10"].value == TRUE:
            if req_t == "کالا":
                status = "جستجوی کالا "
            else:
                status = "جستجوی پیمانکار"
        elif worksheet["X10"].value == TRUE:
            status = "درخواست اطلاعات از سایت"
        else:
            status = "ارجا به بازرگانی"
        # project name
        project_name = worksheet["G6"].value
        # # get the project name and code
        table = chooseTable(project_name)
        code = worksheet["H6"].value
        # check if it needs updates
        if delete_needed:
            query = f"DELETE FROM {table} WHERE req_n=%s"
            value = (req_n,)
            query_main = f"DELETE FROM main WHERE req_n=%s"
            value_main = (req_n,)
            cursor.execute(query, value)
            cursor.execute(query_main, value_main)
        # applicant
        Applicant = worksheet["D6"].value
        # get the materials value
        for row in range(8, 14):
            if req_t == "کالا":
                prod = worksheet.cell(row=row, column=3).value
                qty = worksheet.cell(row=row, column=9).value
                available = worksheet.cell(row=row, column=11).value
                place_of_usage = worksheet.cell(row=row, column=14).value
                unit = worksheet.cell(row=row, column=13).value
            else:
                prod = worksheet.cell(row=row, column=3).value
                place_of_usage = worksheet["K7"].value
                qty = None
                available = None
                unit = None
            if prod != None and prod != "شرح خدمات درخواستی :":
                sql = f"INSERT INTO {table} (product, req_n, Qty, o_date, available_in_stock, ref_date, place_of_usage, unit ,Applicant, st_request, req_t) VALUES (%s, %s, %s,%s,%s, %s, %s, %s, %s, %s, %s)"
                val = (
                    prod,
                    req_n,
                    qty,
                    o_date,
                    available,
                    ref_date,
                    place_of_usage,
                    unit,
                    Applicant,
                    status,
                    req_t,
                )
                cursor.execute(sql, val)
        if status != "invalid":
            sql_main = f"INSERT INTO main (project, req_n, o_date, f_date, p_code, st_request, req_t, location) VALUES (%s ,%s, %s, %s, %s, %s, %s, %s)"
            val_main = (
                project_name,
                req_n,
                o_date,
                ref_date,
                code,
                status,
                req_t,
                file_path,
            )
            cursor.execute(sql_main, val_main)
            conn.commit()
            done_wondow()

    except KeyError as e:
        errorWindow("است قبول قابل غیر فایل نام\n است RE#-#####-###-### قبول قابل فرمت")
    except project_manager:
        errorWindow(req_n + " فایل برای پروژه مدیر تایید عدم ")
def done_wondow():
        new_window = ctk.CTk()
        new_window.title("saved")
        new_window.geometry("300x100")
        done_label = ctk.CTkLabel(
            new_window,
            text="شد ثبت موفقیت با اطلاعات",
        )
        done_label.place(x=100, y=15)
        ok_button = ctk.CTkButton(new_window, text="ok", command=new_window.destroy)
        ok_button.place(x=80, y=60)
        new_window.mainloop()
        
if __name__ == "__main__":
    excel_file_path = sys.argv[1]
    save_data_to_database(True,excel_file_path)
