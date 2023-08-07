import pandas as pd
import pymysql
import openpyxl
import os
import tkinter as tk
from tkinter import *
from tkinter.ttk import *
import tkinter as tk
from tkinter import ttk
import os
conn = pymysql.connect(
        host="localhost", user="root", password="1122", database="request_control"
    )
cursor=conn.cursor()


def errorWindow(error_massage):
    newWindow = Toplevel(window)
    newWindow.title("Error")
    newWindow.geometry("400x100")
    title_label = tk.Label(
        newWindow, text=error_massage, fg="red", font="Verdana 10 bold"
    )
    title_label.place(x=50, y=10)
    close_button = tk.Button(
        newWindow, text="ok", height=1, width=8, command=newWindow.destroy
    )
    close_button.place(x=160, y=60)


def run_program():
    window.destroy()
    os.system("python view.py")


def chooseTable(name):
    return {
        "فاضلاب قم 5 ساله": "quem",
        "فاضلاب خین عرب": "khin",
        "فاضلاب التیمور": "alteymour",
        "آبرسانی جاسک": "jask",
        "رودان 2": "roudan",
        "رشت": "rasht",
        "مرکزی": "markazi",
    }[name]


def chooseCode(name):
    return {
        "فاضلاب قم 5 ساله": 777,
        "فاضلاب خین عرب": 770,
        "فاضلاب التیمور": 880,
        "آبرسانی جاسک": 667,
        "رودان 2": 666,
        "رشت": 210,
        "مرکزی": 110,
    }[name]
    
    
def checkForFile():
    file_path = request_number_entry.get()
    file_name = os.path.basename(file_path)
    file_name= file_name.replace(".xlsm", "")
    query = "SELECT * FROM main WHERE req_n = %s"
    value = (file_name,)  
    cursor.execute(query, value)  
    if cursor.fetchone():
        updateWindow(file_name, file_path)
    else:
        save_data_to_database(file_name, False, file_path)
        
def updateWindow(file_name, file_path):
    newWindow = Toplevel(window)
    newWindow.title("Error")
    newWindow.geometry("400x100")
    title_label = tk.Label(
        newWindow, text="کنید؟ آپدیت را آن میخواهید آیا. دارد وجود دیتابیس در فایل این" , fg="red", font="Verdana 10 bold"
    )
    title_label.place(x=50, y=10)
    def update():
        newWindow.destroy()
        save_data_to_database(file_name, True, file_path)
    update_button = tk.Button( newWindow, text="update", height=1, width=8, command=update)
    update_button.place(x=190, y=60)
    def close():
        newWindow.destroy()
    close_button = tk.Button( newWindow, text="cancle", height=1, width=8, command=close)
    close_button.place(x=110, y=60)
def getStatus(index):
    return{
        9: None,
        10: "درخواست اطلاعات از سایت",
        11: "جستجوی کالا ",
        12: "تأمین بودجه",
        13:""
    }[index]
def save_data_to_database(file_name, delete_needed, file_path):
    try:
        workbook = openpyxl.load_workbook(
            file_path, keep_vba=True, data_only=True
        )

        # file_name= file_name.replace()
        worksheet = workbook[file_name]
        req_n = file_name
        o_date = worksheet.cell(row=6, column=14).value
        ref_date = worksheet.cell(row=17, column=14).value
        if(worksheet.cell(row=22, column=24).value==True):
            status = "توقف"
        else:
            if(worksheet.cell(row=10, column=24).value==FALSE):
                status = "درخواست اطلاعات از سایت"
            if(worksheet.cell(row=11, column=24).value==FALSE):
                status = "جستجوی کالا "
            if(worksheet.cell(row=12, column=24).value==FALSE):
                status = "تأمین بودجه"
            if(worksheet.cell(row=13, column=24).value==FALSE):
                status = "پروسه ی خرید"
            else:
                status = "تکمیل و ارسال به سایت"
        project_name = worksheet["G6"].value
        table = chooseTable(project_name)
        code = chooseCode(project_name)
        if(delete_needed):
            query = f"DELETE FROM {table} WHERE req_n=%s"
            value = (file_name,)
            query_main= f"DELETE FROM main WHERE req_n=%s"
            value_main = (file_name,)
            cursor.execute(query, value) 
            cursor.execute(query_main, value_main) 
        Applicant = worksheet.cell(row=6, column=4).value
        for row in range(8, 14):
            prod = worksheet.cell(row=row, column=3).value
            qty = worksheet.cell(row=row, column=9).value
            available = worksheet.cell(row=row, column=11).value
            place_of_usage = worksheet.cell(row=row, column=14).value
            unit =worksheet.cell(row=row, column=13).value
            if prod != None:
                sql = f"INSERT INTO {table} (product, req_n, Qty, o_date, available_in_stock, ref_date, place_of_usage, unit ,Applicant, st_request) VALUES (%s, %s, %s,%s,%s, %s, %s, %s, %s, %s)"
                val = (prod, req_n, qty, o_date, available, ref_date, place_of_usage, unit, Applicant, status)
                cursor.execute(sql, val)
        done_label = tk.Label(
            window, text="شد ثبت موفقیت با اطلاعات", fg="green", font="Verdana 10 bold"
        )
        done_label.place(x=200, y=200)
        ok_button = tk.Button(window, text="ok", height=1, width=8, command=run_program)
        ok_button.place(x=250, y=250)
        sql_main = f"INSERT INTO main (project, req_n, o_date, f_date, p_code, st_request) VALUES (%s ,%s, %s, %s, %s, %s)"
        val_main = (project_name, req_n, o_date, ref_date, code, status)
        cursor.execute(sql_main, val_main)
        conn.commit()
        cursor.close()
        conn.close()
    except FileNotFoundError as e:
        errorWindow(
            "ندارد وجود پوشه در نظر مورد فایل\nباشد REM-#####-###-### صورت به باید نام فرمت"
        )
    # except KeyError as e:
    #     errorWindow(
    #         "است قبول قابل غیر فایل نام\n آورید در REM-#####-###-### صورت به را آن لطفا", "ok"
    #     )


# Create the main window
window = tk.Tk()
window.geometry("500x500")
window.title("سیستم کنرل درخواست کالا و کار")
title_label = tk.Label(window, text="شرکت نیلآب صنعت البرز")
title_label.place(x=190, y=10)

# Create the project name label and entry widgets
request_number_label = tk.Label(window, text="Riquest Number:")
request_number_label.place(x=10, y=80)

request_number_entry = tk.Entry(window)
request_number_entry.place(x=105, y=80)

# Create a search button for the project name
request_number_button = tk.Button(
    window, text="Submite", command=lambda: checkForFile()
)
request_number_button.place(x=250, y=75)

back_button = tk.Button(
    window, text="  back  ", command=lambda: run_program()
)
back_button.place(x=320, y=75)
# Display the company logo image
image_label = tk.Label(window)
image_file = tk.PhotoImage(file="C:/Users/alire/Desktop/rominas workspace/logo5.png")
resized_image_file = image_file.subsample(3, 3)

# Set the image for the label
image_label.config(image=resized_image_file)
image_label.place(x=400, y=10)
# Run the GUI
window.mainloop()
