import os
import openpyxl
import pymysql
import tkinter as tk
from tkinter import *
from tkinter.ttk import *
import tkinter as tk
from tkinter import ttk
import os
from tkinter import filedialog
from openpyxl.styles import PatternFill
import pyperclip

conn = pymysql.connect(
    host="localhost", user="root", password="1122", database="request_control"
)
cursor = conn.cursor()


def notFound():
    newWindow = Toplevel(window)
    newWindow.title("FileNotFound")
    newWindow.geometry("300x100")
    title_label = tk.Label(
        newWindow,
        text="ندارد وجود درخواست ",
        fg="red",
        font="Verdana 10 bold",
    )
    title_label.place(x=100, y=10)

    def ok():
        newWindow.destroy()

    update_button = tk.Button(newWindow, text="ok", height=1, width=8, command=ok)
    update_button.place(x=120, y=60)


def checkForFile(req_n, table_name):
    if len(req_n) == 0:
        querynull = f"SELECT * FROM {table_name}"
        cursor.execute(querynull)
    else:
        query = f"SELECT * FROM {table_name} WHERE req_n = %s"
        value = (req_n,)
        cursor.execute(query, value)
    data = cursor.fetchall()
    if data:

        def print():
            printData(list(data), table_name)

        back_button = tk.Button(window, text="   print   ", command=print)
        back_button.place(x=1000, y=165)
        if table_name == "main":
            mainTable(data)
            mainFiltering(data)

        else:
            showTable(data)
    else:
        notFound()


def mainFiltering(data):
    project_name = tk.StringVar()
    options = ttk.Combobox(window, width=17, textvariable=project_name)
    options.place(x=440, y=60)
    options["values"] = (
        "مرکزی",
        "رشت",
        "رودان 2",
        "آبرسانی جاسک",
        "فاضلاب التیمور",
        "فاضلاب خین عرب",
        "فاضلاب قم 5 ساله",
    )

    def filtering():
        table_name = project_name.get()
        query = f"SELECT * FROM main WHERE project = %s"
        value = (table_name,)
        cursor.execute(query, value)
        data = cursor.fetchall()
        mainTable(data)

    button = tk.Button(window, text="Search in main", command=filtering)
    button.place(x=590, y=55)
    return data


def printData(data, table_name):
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if table_name == "main":
        worksheet.cell(row=1, column=1).value = "نام پروژه"
        worksheet.cell(row=1, column=2).value = "کد پروژه"
        worksheet.cell(row=1, column=3).value = "شماره درخواست"
        worksheet.cell(row=1, column=4).value = "تاریخ درخواست"
        worksheet.cell(row=1, column=5).value = "تاریخ ارجا"
        worksheet.cell(row=1, column=6).value = "نوع درخواست"
        worksheet.cell(row=1, column=7).value = "وضعیت درخواست"
    else:
        worksheet.cell(row=1, column=1).value = "درخواست"
        worksheet.cell(row=1, column=2).value = "شماره درخواست"
        worksheet.cell(row=1, column=3).value = "تاریخ درخواست"
        worksheet.cell(row=1, column=4).value = "تاریخ ارجا"
        worksheet.cell(row=1, column=5).value = "تعداد"
        worksheet.cell(row=1, column=6).value = "موجود"
        worksheet.cell(row=1, column=7).value = "واحد"
        worksheet.cell(row=1, column=8).value = "محل مصرف"
        worksheet.cell(row=1, column=11).value = "نوع درخواست"
        worksheet.cell(row=1, column=9).value = "درخواست کننده"
        worksheet.cell(row=1, column=10).value = "وضعیت درخواست"

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = cell_value
            if cell_value == "توقف":
                # Apply red fill to the entire row
                fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                for cell in worksheet[row_idx]:
                    cell.fill = fill
            elif cell_value == "تکمیل و ارسال به سایت":
                # Apply red fill to the entire row
                fill = PatternFill(
                    start_color="00FF50", end_color="00FF50", fill_type="solid"
                )
                for cell in worksheet[row_idx]:
                    cell.fill = fill
    for column in range(ord("A"), ord("K")):
        column_letter = chr(column)
        worksheet.column_dimensions[column_letter].width = 18

    workbook.save(file_path)

    newWindow = Toplevel(window)
    newWindow.geometry("200x100")
    title_label = tk.Label(
        newWindow, text="شد ثبت موفقیت با اطلاعات", fg="green", font="Verdana 10 bold"
    )
    title_label.place(x=50, y=10)
    close_button = tk.Button(
        newWindow, text="ok", height=1, width=8, command=newWindow.destroy
    )
    close_button.place(x=80, y=60)

def copy_to_clipboard(event, table):
    selected_row = table.focus()
    data = table.item(selected_row)["values"]
    # Copy req_num to clipboard
    pyperclip.copy(data[2])

def handle_double_click(event, table):
    # Get the selected row
    selected_row = table.focus()
    # Perform the desired action
    # For example, print the selected row's data
    data = table.item(selected_row)["values"]
    req_n = data[3]

    newWindow = Toplevel(window)
    newWindow.geometry("300x100")
    title_label = tk.Label(
        newWindow, text="پرداخت دستور صدور", fg="green", font="Verdana 10 bold"
    )
    title_label.place(x=100, y=10)

    def insert():
        workbook = openpyxl.load_workbook(
            "C:/Users/alire/Desktop/rominas workspace/payment order.xlsx"
        )
        sheet = workbook["Sheet1"]
        # file = open('C:/Users/alire/Desktop/rominas workspace/logo5.png', 'rb')
        # data = BytesIO(file.read())
        # file.close()
        # sheet.insert_image('C3', 'C:/Users/alire/Desktop/rominas workspace/logo5.png', {'image_data': data})
        sheet["F6"] = req_n
        workbook.save(
            filedialog.asksaveasfilename(
                defaultextension=".xlsx", initialfile=req_n + "-PO"
            )
        )
        newWindow.destroy()

    insert_button = tk.Button(
        newWindow, text="insert", height=1, width=8, command=insert
    )
    insert_button.place(x=80, y=60)

    close_button = tk.Button(
        newWindow, text="cancle", height=1, width=8, command=newWindow.destroy
    )
    close_button.place(x=160, y=60)


def mainTable(data):
    table = ttk.Treeview(
        window,
        columns=("1", "2", "3", "4", "5", "6", "7", "8", "9", "10"),
        show="headings",
        height=10,
    )
    table.pack()
    table.column("1", anchor=CENTER, stretch=YES, width=40)
    table.heading("1", text="Id")
    table.column("2", anchor=CENTER, stretch=YES, width=100)
    table.heading("2", text="نام پروژه")
    table.column("3", anchor=CENTER, stretch=YES, width=60)
    table.heading("3", text="کد پروژه")
    table.column("4", anchor=CENTER, stretch=YES, width=120)
    table.heading("4", text="شماره درخواست")
    table.column("5", anchor=CENTER, stretch=YES, width=100)
    table.heading("5", text="تاریخ درخواست")
    table.column("6", anchor=CENTER, stretch=YES, width=100)
    table.heading("6", text="تاریخ ارجا")
    table.column("7", anchor=CENTER, stretch=YES, width=100)
    table.heading("7", text="نوع درخواست")
    table.column("8", anchor=CENTER, stretch=YES, width=140)
    table.heading("8", text="وضعیت درخواست")
    table.column("9", anchor=CENTER, stretch=YES, width=90)
    table.heading("9", text="پرداخت 1")
    table.column("10", anchor=CENTER, stretch=YES, width=90)
    table.heading("10", text="پرداخت 2")
    table.place(x=50, y=170)
    i = 0
    for row in data:
        table.insert(
            "",
            "end",
            text=i,
            values=(
                i + 1,
                data[i][0],
                data[i][1],
                data[i][2],
                data[i][3],
                data[i][4],
                data[i][5],
                data[i][6],
            ),
        )
        i += 1
    table.bind("<Double-1>", lambda event: handle_double_click(event, table))


def showTable(data):
    table = ttk.Treeview(
        window,
        columns=("1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"),
        show="headings",
        height=10,
    )
    table.pack()
    table.column("1", anchor=CENTER, stretch=YES, width=10)
    table.heading("1", text="Id")
    table.column("2", anchor=CENTER, stretch=YES, width=120)
    table.heading("2", text="درخواست")
    table.column("3", anchor=CENTER, stretch=YES, width=120)
    table.heading("3", text="شماره درخواست")
    table.column("4", anchor=CENTER, stretch=YES, width=90)
    table.heading("4", text="تاریخ درخواست")
    table.column("5", anchor=CENTER, stretch=YES, width=90)
    table.heading("5", text="تاریخ ارجا")
    table.column("10", anchor=CENTER, stretch=YES, width=50)
    table.heading("10", text="درخواست کننده")
    table.column("6", anchor=CENTER, stretch=YES, width=50)
    table.heading("6", text="تعداد")
    table.column("7", anchor=CENTER, stretch=YES, width=50)
    table.heading("7", text="موجود")
    table.column("9", anchor=CENTER, stretch=YES, width=100)
    table.heading("9", text="محل مصرف")
    table.column("8", anchor=CENTER, stretch=YES, width=50)
    table.heading("8", text="واحد")
    table.column("11", anchor=CENTER, stretch=YES, width=120)
    table.heading("11", text="وضعیت درخواست")
    table.column("12", anchor=CENTER, stretch=YES, width=90)
    table.heading("12", text="نوع درخواست")
    table.place(x=50, y=170)
    i = 0
    for row in data:
        table.insert(
            "",
            "end",
            text=i,
            values=(
                i + 1,
                data[i][0],
                data[i][1],
                data[i][2],
                data[i][3],
                data[i][4],
                data[i][5],
                data[i][6],
                data[i][7],
                data[i][8],
                data[i][9],
                data[i][10],
            ),
            tag=("odd"),
        )
        i += 1
    table.bind("<Double-1>", lambda event: copy_to_clipboard(event, table))


def upload():
    window.destroy()
    os.system("python pythontest.py")


def create_upload_button(window):
    upload_button = tk.Button(window, text="Upload New File", command=upload)
    upload_button.place(x=50, y=100)


def chooseTable(name):
    return {
        "فاضلاب قم 5 ساله": "quem",
        "فاضلاب خین عرب": "khin",
        "فاضلاب التیمور": "alteymour",
        "آبرسانی جاسک": "jask",
        "رودان 2": "roudan",
        "رشت": "rasht",
        "مرکزی": "markazi",
        "main": "main",
    }[name]


def discriptions():
    table = ttk.Treeview(window, columns=("1", "2"), show="headings", height=6)
    table.pack()
    table.column("1", anchor=CENTER, stretch=YES, width=100)
    table.heading("1", text="نام پروژه")
    table.column("2", anchor=CENTER, stretch=YES, width=50)
    table.heading("2", text="کد پروژه")
    table.insert("", "end", values=("فاضلاب قم 5 ساله", "777"))
    table.insert("", "end", values=("فاضلاب خین عرب", "770"))
    table.insert("", "end", values=("فاضلاب التیمور", "880"))
    table.insert("", "end", values=("آبرسانی جاسک", "667"))
    table.insert("", "end", values=("رودان 2", "666"))
    table.insert("", "end", values=("رشت", "210"))
    table.insert("", "end", values=("مرکزی", "110"))
    table.place(x=839, y=5)
    


# Create the main window
window = tk.Tk()
window.geometry("1200x500")
window.title("سیستم کنرل درخواست کالا و کار")
#####################################################
request_number_label = tk.Label(window, text="Request Number:")
request_number_label.place(x=50, y=20)
# Create the request number text box
request_number_entry = tk.Entry(window)
request_number_entry.place(x=160, y=20)


#####################################################
# Create a label for the project name choice box
project_name_label = tk.Label(window, text="Project Name:")
project_name_label.place(x=51, y=60)

# Create the project name choice box
project_name_var = tk.StringVar()
options = ttk.Combobox(window, width=17, textvariable=project_name_var)
options.place(x=160, y=60)
options["values"] = (
    "مرکزی",
    "رشت",
    "رودان 2",
    "آبرسانی جاسک",
    "فاضلاب التیمور",
    "فاضلاب خین عرب",
    "فاضلاب قم 5 ساله",
    "main",
)


def search_project():
    selected_value = project_name_var.get()
    checkForFile(request_number_entry.get(), chooseTable(selected_value))


project_name_button = tk.Button(window, text="Search", command=search_project)
project_name_button.place(x=350, y=55)


###################################################
create_upload_button(window)
image_label = tk.Label(window)
image_file = tk.PhotoImage(file="C:/Users/alire/Desktop/rominas workspace/logo5.png")
resized_image_file = image_file.subsample(2, 2)
# Set the image for the label
image_label.config(image=resized_image_file)
image_label.place(x=1050, y=10)
discriptions()
# Run the GUI
window.mainloop()
