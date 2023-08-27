import pandas as pd
import pymysql
import openpyxl
import os
import tkinter as tk
from tkinter import *
from tkinter.ttk import *
import tkinter as tk
from tkinter import ttk
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename

conn = pymysql.connect(
    host="localhost", user="root", password="1122", database="request_control"
)
cursor = conn.cursor()


def errorWindow(error_massage):
    newWindow = Toplevel(window)
    newWindow.title("Error")
    newWindow.geometry("400x100")
    title_label = tk.Label(
        newWindow, text=error_massage, fg="red", font="Verdana 10 bold"
    )
    title_label.place(x=50, y=10)
    close_button = tk.Button(
        newWindow, text="ok", height=1, width=8, command=newWindow.destroy
    )
    close_button.place(x=160, y=60)


def run_program():
    window.destroy()
    os.system("python RCS.py")


def chooseTable(name):
    return {
        "فاضلاب قم 5 ساله": "quem",
        "فاضلاب خین عرب": "khin",
        "فاضلاب التیمور": "alteymour",
        "آبرسانی جاسک": "jask",
        "رودان 2": "roudan",
        "رشت": "rasht",
        "مرکزی": "markazi",
    }[name]


def chooseCode(name):
    return {
        "فاضلاب قم 5 ساله": 777,
        "فاضلاب خین عرب": 770,
        "فاضلاب التیمور": 880,
        "آبرسانی جاسک": 667,
        "رودان 2": 666,
        "رشت": 210,
        "مرکزی": 110,
    }[name]


def checkForFile(fileLocation):
    file_name = os.path.basename(fileLocation)
    file_name = file_name.replace(".xlsm", "")
    query = "SELECT * FROM main WHERE req_n = %s"
    value = (file_name,)
    cursor.execute(query, value)
    if cursor.fetchone():
        updateWindow(file_name, fileLocation)
    else:
        save_data_to_database(file_name, False, fileLocation)


def updateWindow(file_name, file_path):
    newWindow = Toplevel(window)
    newWindow.title("Error")
    newWindow.geometry("400x100")
    title_label = tk.Label(
        newWindow,
        text="کنید؟ آپدیت را آن میخواهید آیا. دارد وجود دیتابیس در فایل این",
        fg="red",
        font="Verdana 10 bold",
    )
    title_label.place(x=50, y=10)

    def update():
        newWindow.destroy()
        save_data_to_database(file_name, True, file_path)

    update_button = tk.Button(
        newWindow, text="update", height=1, width=8, command=update
    )
    update_button.place(x=190, y=60)

    def close():
        newWindow.destroy()

    close_button = tk.Button(newWindow, text="cancle", height=1, width=8, command=close)
    close_button.place(x=110, y=60)


def getStatus(index):
    return {
        9: None,
        10: "درخواست اطلاعات از سایت",
        11: "جستجوی کالا ",
        12: "تأمین بودجه",
        13: "",
    }[index]


def save_data_to_database(file_name, delete_needed, file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, keep_vba=True, data_only=True)

        
        worksheet = workbook[file_name]
        # file name
        req_n = file_name
        # request type
        if "REM" in req_n:
            req_t = "کالا"
        else:
            req_t = "کار"
        # order date
        o_date = worksheet["N6"].value
        # referral date
        if req_t == "کالا":
            ref_date = worksheet["N17"].value
        else:
            ref_date = worksheet["N16"].value
        #request status
        if worksheet["X8"].value == True:
            status = "توقف"
        else:
            if worksheet["X13"].value == TRUE:
                if req_t == "کالا":
                    status = "تکمیل و ارسال به سایت"
                else:
                    status = "تکمیل کار"
            elif worksheet["X12"].value == TRUE:
                if req_t == "کالا":
                    status = "پروسه ی خرید"
                else:
                    status = "پروسه ی ارجاع کار به پیمانکار"

            elif worksheet["X11"].value == TRUE:
                status = "تأمین بودجه"
            elif worksheet["X10"].value == TRUE:
                if req_t == "کالا":
                    status = "جستجوی کالا "
                else:
                    status = "جستجوی پیمانکار"
            else:
                status = "درخواست اطلاعات از سایت"
        #project name
        project_name = worksheet["G6"].value
        
        #get the project name and code
        table = chooseTable(project_name)
        code = chooseCode(project_name)
        
        #check if it needs updates
        if delete_needed:
            query = f"DELETE FROM {table} WHERE req_n=%s"
            value = (file_name,)
            query_main = f"DELETE FROM main WHERE req_n=%s"
            value_main = (file_name,)
            cursor.execute(query, value)
            cursor.execute(query_main, value_main)
        #applicant
        Applicant = worksheet["D6"].value
        #get the materials value
        for row in range(8, 14):
            if req_t == "کالا":
                prod = worksheet.cell(row=row, column=3).value
                qty = worksheet.cell(row=row, column=9).value
                available = worksheet.cell(row=row, column=11).value
                place_of_usage = worksheet.cell(row=row, column=14).value
                unit = worksheet.cell(row=row, column=13).value
            else:
                row+=1
                prod = worksheet.cell(row=row, column=3).value
                place_of_usage = worksheet["K7"].value
                qty = None
                available = None
                unit = None
            if prod != None and prod != "شرح خدمات درخواستی :	":
                sql = f"INSERT INTO {table} (product, req_n, Qty, o_date, available_in_stock, ref_date, place_of_usage, unit ,Applicant, st_request, req_t) VALUES (%s, %s, %s,%s,%s, %s, %s, %s, %s, %s, %s)"
                val = (
                    prod,
                    req_n,
                    qty,
                    o_date,
                    available,
                    ref_date,
                    place_of_usage,
                    unit,
                    Applicant,
                    status,
                    req_t,
                )
                cursor.execute(sql, val)
        done_label = tk.Label(
            window, text="شد ثبت موفقیت با اطلاعات", fg="green", font="Verdana 10 bold"
        )
        done_label.place(x=200, y=200)
        ok_button = tk.Button(window, text="ok", height=1, width=8, command=run_program)
        ok_button.place(x=250, y=250)
        sql_main = f"INSERT INTO main (project, req_n, o_date, f_date, p_code, st_request, req_t) VALUES (%s ,%s, %s, %s, %s, %s, %s)"
        val_main = (project_name, req_n, o_date, ref_date, code, status, req_t)
        cursor.execute(sql_main, val_main)
        conn.commit()
    except FileNotFoundError as e:
        errorWindow(
            "ندارد وجود پوشه در نظر مورد فایل\nباشد RE#-#####-###-### صورت به باید نام فرمت"
        )
    except KeyError as e:
        errorWindow(
            "است قبول قابل غیر فایل نام\n است RE#-#####-###-### قبول قابل فرمت"
        )


def codeTableDis():
    table = ttk.Treeview(window, columns=("1", "2"), show="headings", height=6)
    table.pack()
    table.column("1", anchor=CENTER, stretch=YES, width=100)
    table.heading("1", text="نام پروژه")
    table.column("2", anchor=CENTER, stretch=YES, width=50)
    table.heading("2", text="کد پروژه")
    table.insert("", "end", values=("فاضلاب قم 5 ساله", "777"))
    table.insert("", "end", values=("فاضلاب خین عرب", "770"))
    table.insert("", "end", values=("فاضلاب التیمور", "880"))
    table.insert("", "end", values=("آبرسانی جاسک", "667"))
    table.insert("", "end", values=("رودان 2", "666"))
    table.insert("", "end", values=("رشت", "210"))
    table.insert("", "end", values=("مرکزی", "110"))
    table.place(x=250, y=80)


# Create the main window
window = tk.Tk()
window.geometry("500x500")
window.title("سیستم کنرل درخواست کالا و کار")
title_label = tk.Label(window, text="شرکت نیلآب صنعت البرز")
title_label.place(x=190, y=10)

# Create the project name label and entry widgets
request_number_label = tk.Label(window, text="Select File : ")
request_number_label.place(x=10, y=80)


def askForFile():
    checkForFile(askopenfilename())


find_button = tk.Button(window, text="  bowers   ", command=askForFile)
find_button.place(x=90, y=78)

back_button = tk.Button(window, text="   back   ", command=lambda: run_program())
back_button.place(x=170, y=78)
# Display the company logo image
image_label = tk.Label(window)
image_file = tk.PhotoImage(file="C:/Users/alire/Desktop/rominas workspace/logo5.png")
resized_image_file = image_file.subsample(3, 3)

# Set the image for the label
image_label.config(image=resized_image_file)
image_label.place(x=400, y=10)
codeTableDis()
# Run the GUI
window.mainloop()
cursor.close()
conn.close()
